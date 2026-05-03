"""
Torgor & Tweed — QA Progress Tracker Generator
Run: python generate_qa_tracker.py
Output: Torgor_and_Tweed_QA_Tracker.xlsx (in the same folder)
Requires: pip install openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import datetime

OUTPUT = "Torgor_and_Tweed_QA_Tracker.xlsx"
QA_DATE = "2026-05-03"

# ── BRAND COLORS ──────────────────────────────────────────────────────────────
BRAND_BLACK  = "1A1A1A"
BRAND_GOLD   = "C8A96E"
WHITE        = "FFFFFF"
LIGHT_GOLD   = "F5EFE6"
PASS_BG      = "D4EDDA"; PASS_FG    = "155724"
FAIL_BG      = "F8D7DA"; FAIL_FG    = "721C24"
BLOCKED_BG   = "FFF3CD"; BLOCKED_FG = "856404"
NOTRUN_BG    = "E2E3E5"; NOTRUN_FG  = "383D41"
NOTPLAN_BG   = "D6D8DB"; NOTPLAN_FG = "555555"
CRIT_COLOR   = "FF4444"
HIGH_COLOR   = "FF8800"
MED_COLOR    = "0066CC"
LOW_COLOR    = "28A745"
GRAY_ROW1    = "FAFAFA"
GRAY_ROW2    = "FFFFFF"

# ── HELPER STYLES ─────────────────────────────────────────────────────────────
def font(name="Arial", size=10, bold=False, color="000000", italic=False):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)

def fill(color):
    return PatternFill("solid", fgColor=color)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(color="D0D0D0"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border_bottom(color=BRAND_GOLD):
    m = Side(style="medium", color=color)
    t = Side(style="thin", color="444444")
    return Border(top=t, bottom=m)

def col_letter(n):
    return get_column_letter(n)

STATUS_STYLE = {
    "PASS":        (PASS_BG,    PASS_FG),
    "FAIL":        (FAIL_BG,    FAIL_FG),
    "BLOCKED":     (BLOCKED_BG, BLOCKED_FG),
    "NOT RUN":     (NOTRUN_BG,  NOTRUN_FG),
    "NOT PLANNED": (NOTPLAN_BG, NOTPLAN_FG),
}

PRIO_BG = {
    "Critical": "FFCCCC",
    "High":     "FFE5CC",
    "Medium":   "CCE5FF",
    "Low":      "CCFFCC",
}

# ── QA TEST DATA ──────────────────────────────────────────────────────────────
# [ID, Category, Scenario, Steps, Expected, Status, Priority, Notes]
TESTS = [
    # Functional
    ("TC-001","Functional","Product Search","Enter keyword in search bar","Relevant products displayed","BLOCKED","High","Application not yet built — no src/ files exist in project"),
    ("TC-002","Functional","Advanced Filters","Apply category, price, rating filters","Correct filtered results shown","BLOCKED","High","Application not yet built"),
    ("TC-003","Functional","Product Sorting","Sort by price low→high","Products sorted correctly","BLOCKED","Medium","Application not yet built"),
    ("TC-004","Functional","Product Detail Page","Open product page","Images, description, reviews load","BLOCKED","High","Application not yet built"),
    ("TC-005","Functional","Add to Cart","Add product to cart","Product appears with correct price","BLOCKED","Critical","Application not yet built"),
    ("TC-006","Functional","Remove from Cart","Remove product","Product removed, total updated","BLOCKED","Critical","Application not yet built"),
    ("TC-007","Functional","Update Quantity","Change quantity in cart","Total updates correctly","BLOCKED","High","Application not yet built"),
    ("TC-008","Functional","Guest Checkout","Checkout without login","Order placed, confirmation email sent","BLOCKED","Critical","Application not yet built"),
    ("TC-009","Functional","Registered Checkout","Checkout logged in","Order placed, visible in account","BLOCKED","Critical","Application not yet built"),
    ("TC-010","Functional","Multiple Payment Methods","Pay via COD, bKash, Nagad","Payment processed successfully","BLOCKED","Critical","Payment gateway integration not yet implemented"),
    ("TC-011","Functional","Order Confirmation Email","Place order","Email received with details","BLOCKED","High","Resend SDK not yet wired — RESEND_API_KEY empty in .env.example"),
    ("TC-012","Functional","Order Tracking","Check My Orders","Status updates correctly","BLOCKED","High","Application not yet built"),
    # Auth
    ("TC-013","Auth","User Registration","Sign up with valid details","Account created, welcome email sent","BLOCKED","High","NextAuth not yet configured"),
    ("TC-014","Auth","Login","Enter valid credentials","User logged in","BLOCKED","High","NextAuth not yet configured"),
    ("TC-015","Auth","Logout","Click logout","User logged out","BLOCKED","Medium","NextAuth not yet configured"),
    ("TC-016","Auth","Password Reset","Request reset","Reset link received","BLOCKED","Medium","Resend SDK + auth flows not yet implemented"),
    # Functional (cont)
    ("TC-017","Functional","Wishlist","Add product to wishlist","Product saved in wishlist","BLOCKED","Medium","Application not yet built"),
    ("TC-018","Functional","Compare Products","Select multiple products","Comparison table displayed","BLOCKED","Low","Application not yet built"),
    ("TC-019","Functional","Review Submission","Submit product review","Review posted successfully","BLOCKED","Medium","Application not yet built"),
    ("TC-020","Functional","Coupon Code","Apply discount code","Discount applied correctly","BLOCKED","Medium","Application not yet built"),
    # UI/UX
    ("TC-021","UI/UX","Navigation","Browse menus","Clear, consistent navigation","BLOCKED","High","No UI built yet"),
    ("TC-022","UI/UX","Mobile Responsiveness","Open site on mobile","Layout adjusts properly","BLOCKED","High","No UI built; Tailwind CSS configured in postcss.config.js ✓"),
    ("TC-023","UI/UX","Tablet Responsiveness","Open site on tablet","Layout adjusts properly","BLOCKED","Medium","No UI built yet"),
    ("TC-024","UI/UX","Accessibility – Keyboard","Navigate with keyboard","All elements accessible","BLOCKED","High","No UI built yet"),
    ("TC-025","UI/UX","Accessibility – Alt Text","Inspect product images","Alt text present","BLOCKED","High","No UI built yet"),
    ("TC-026","UI/UX","Form Validation","Submit empty form","Error messages displayed","BLOCKED","High","Zod planned in stack; no forms built yet"),
    ("TC-027","UI/UX","Error Handling","Enter invalid coupon","Clear error shown","BLOCKED","Medium","No UI built yet"),
    ("TC-028","UI/UX","Multi-language Support","Switch language","Content translated correctly","NOT PLANNED","Low","Not in PLAN.md scope — single market Bangladesh (Bengali/English)"),
    # Performance
    ("TC-029","Performance","Homepage Load Speed","Open homepage","Loads <3 seconds","BLOCKED","High","Site not deployed"),
    ("TC-030","Performance","Product Page Speed","Open product page","Loads <3 seconds","BLOCKED","High","Site not deployed"),
    ("TC-031","Performance","Checkout Speed","Proceed to checkout","Loads <5 seconds","BLOCKED","Medium","Site not deployed"),
    ("TC-032","Performance","Stress Test","Simulate 1000 users","Site remains stable","BLOCKED","Medium","Site not deployed"),
    ("TC-033","Performance","Database Query Efficiency","Perform search","Queries optimized","BLOCKED","Medium","No DB schema built yet (schema.prisma missing)"),
    ("TC-034","Performance","CDN Performance","Load images","Images served via CDN","BLOCKED","Medium","Cloudinary config present in .env.example; not connected yet"),
    # Security
    ("TC-035","Security","SQL Injection","Enter ' OR 1=1 --","Input sanitized","BLOCKED","Critical","No forms exist; Prisma ORM planned (protects by default) ✓ design direction"),
    ("TC-036","Security","XSS","Enter <script> in form","Script blocked","BLOCKED","Critical","No forms exist; Next.js escapes output by default ✓ design direction"),
    ("TC-037","Security","CSRF","Submit unauthorized request","Request rejected","BLOCKED","Critical","NextAuth v5 provides CSRF protection ✓ design; not yet implemented"),
    ("TC-038","Security","HTTPS","Access site","Secure HTTPS enforced","BLOCKED","Critical","Site not deployed; ⚠ NEXTAUTH_URL=http://localhost in .env.example — must be https:// in prod"),
    ("TC-039","Security","Secure Cookies","Inspect cookies","Secure/HttpOnly flags set","BLOCKED","Critical","Site not deployed; NextAuth sets HttpOnly by default ✓ design direction"),
    ("TC-040","Security","Password Encryption","Register user","Password stored encrypted","BLOCKED","Critical","bcryptjs in node_modules ✓; NextAuth + PrismaAdapter planned; no auth code yet"),
    ("TC-041","Security","Role-based Access","Try admin page as user","Access denied","BLOCKED","High","No route protection implemented yet"),
    # Browser
    ("TC-042","Browser","Chrome Browser","Open site","Functions correctly","BLOCKED","High","Site not deployed"),
    ("TC-043","Browser","Firefox Browser","Open site","Functions correctly","BLOCKED","High","Site not deployed"),
    ("TC-044","Browser","Safari Browser","Open site","Functions correctly","BLOCKED","Medium","Site not deployed"),
    ("TC-045","Browser","Edge Browser","Open site","Functions correctly","BLOCKED","Medium","Site not deployed"),
    # Device
    ("TC-046","Device","iOS Device","Open site","Functions correctly","BLOCKED","High","Site not deployed"),
    ("TC-047","Device","Android Device","Open site","Functions correctly","BLOCKED","High","Site not deployed"),
    ("TC-048","Device","Windows Desktop","Open site","Functions correctly","BLOCKED","Medium","Site not deployed"),
    ("TC-049","Device","macOS Desktop","Open site","Functions correctly","BLOCKED","Medium","Site not deployed"),
    # Regression
    ("TC-050","Regression","Regression – Cart","Retest cart after update","Works correctly","NOT RUN","Medium","No baseline yet; run after TC-005 through TC-007 pass"),
    ("TC-051","Regression","Regression – Checkout","Retest checkout after update","Works correctly","NOT RUN","Medium","No baseline yet; run after TC-008 through TC-009 pass"),
    # Compliance
    ("TC-052","Compliance","GDPR Compliance","Request data deletion","Data deleted","BLOCKED","High","No user data model built yet; Prisma schema pending"),
    ("TC-053","Compliance","CCPA Compliance","Request opt-out","Opt-out processed","BLOCKED","Medium","Low priority — target market is Bangladesh; not CCPA jurisdiction"),
    ("TC-054","Compliance","PCI DSS Compliance","Inspect payment flow","Meets PCI DSS","BLOCKED","Critical","No payment code yet; COD+bKash+Nagad avoids raw card data ✓ design direction"),
]

CAT_ORDER = ["Functional","Auth","UI/UX","Performance","Security","Browser","Device","Regression","Compliance"]

def count_by(field, value):
    return sum(1 for t in TESTS if t[field] == value)

# ─────────────────────────────────────────────────────────────────────────────
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard")

    # ── Title ──
    ws.merge_cells("A1:G1")
    ws["A1"] = "TORGOR & TWEED — QA PROGRESS REPORT"
    ws["A1"].font = font("Arial", 16, True, WHITE)
    ws["A1"].fill = fill(BRAND_BLACK)
    ws["A1"].alignment = align("center", "center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:G2")
    ws["A2"] = f"QA Date: {QA_DATE}   |   Phase: Pre-Launch (Scaffolding)   |   Assessed: {len(TESTS)} test cases"
    ws["A2"].font = font("Arial", 10, False, "555555", italic=True)
    ws["A2"].fill = fill(LIGHT_GOLD)
    ws["A2"].alignment = align("center", "center")
    ws.row_dimensions[2].height = 22

    ws.row_dimensions[3].height = 10

    # ── Summary section headers ──
    sections = [("A4:B4","RESULTS SUMMARY"), ("D4:E4","PRIORITY BREAKDOWN"), ("G4:G4","STATUS")]
    for rng, label in sections:
        ws.merge_cells(rng) if ":" in rng else None
        c = ws[rng.split(":")[0]]
        c.value = label
        c.font = font("Arial", 10, True, WHITE)
        c.fill = fill(BRAND_GOLD)
        c.alignment = align("center", "center")
    ws.row_dimensions[4].height = 22

    # ── Stats ──
    total    = len(TESTS)
    blocked  = count_by(5, "BLOCKED")
    notrun   = count_by(5, "NOT RUN")
    notplan  = count_by(5, "NOT PLANNED")
    passed   = count_by(5, "PASS")
    failed   = count_by(5, "FAIL")
    pct      = round(passed / total * 100) if total else 0

    prios = {}
    for t in TESTS:
        prios[t[6]] = prios.get(t[6], 0) + 1

    summary_rows = [
        ("Total Tests", total),
        ("Passed",      f"{passed} ({pct}%)"),
        ("Failed",      failed),
        ("Blocked",     blocked),
        ("Not Run",     notrun),
        ("Not Planned", notplan),
    ]
    prio_rows = [
        ("Critical", prios.get("Critical",0)),
        ("High",     prios.get("High",0)),
        ("Medium",   prios.get("Medium",0)),
        ("Low",      prios.get("Low",0)),
    ]

    for i, (label, val) in enumerate(summary_rows):
        r = 5 + i
        ws.row_dimensions[r].height = 20
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = font("Arial", 10, True, BRAND_BLACK)
        ws[f"A{r}"].fill = fill(GRAY_ROW1 if i%2==0 else GRAY_ROW2)
        ws[f"A{r}"].alignment = align("left","center")
        ws[f"A{r}"].border = thin_border()
        ws[f"B{r}"] = val
        ws[f"B{r}"].font = font("Arial", 11, True, BRAND_BLACK)
        ws[f"B{r}"].fill = fill(GRAY_ROW1 if i%2==0 else GRAY_ROW2)
        ws[f"B{r}"].alignment = align("center","center")
        ws[f"B{r}"].border = thin_border()

    for i, (label, val) in enumerate(prio_rows):
        r = 5 + i
        color_map = {"Critical":CRIT_COLOR,"High":HIGH_COLOR,"Medium":MED_COLOR,"Low":LOW_COLOR}
        fc = color_map.get(label, BRAND_BLACK)
        ws[f"D{r}"] = label
        ws[f"D{r}"].font = font("Arial", 10, True, fc)
        ws[f"D{r}"].alignment = align("left","center")
        ws[f"D{r}"].border = thin_border()
        ws[f"E{r}"] = val
        ws[f"E{r}"].font = font("Arial", 11, True, fc)
        ws[f"E{r}"].alignment = align("center","center")
        ws[f"E{r}"].border = thin_border()

    # ── Spacer ──
    ws.row_dimensions[11].height = 14

    # ── Category breakdown header ──
    cat_hdr = ["Category","Total","Pass","Fail","Blocked","Not Run","Status"]
    for ci, h in enumerate(cat_hdr, 1):
        cell = ws.cell(row=12, column=ci, value=h)
        cell.font = font("Arial", 10, True, WHITE)
        cell.fill = fill(BRAND_BLACK)
        cell.alignment = align("center","center")
        cell.border = thin_border()
    ws.row_dimensions[12].height = 22

    for i, cat in enumerate(CAT_ORDER):
        ct    = [t for t in TESTS if t[1] == cat]
        if not ct: continue
        p     = sum(1 for t in ct if t[5]=="PASS")
        f     = sum(1 for t in ct if t[5]=="FAIL")
        b     = sum(1 for t in ct if t[5]=="BLOCKED")
        nr    = sum(1 for t in ct if t[5]=="NOT RUN")
        np_   = sum(1 for t in ct if t[5]=="NOT PLANNED")
        status_lbl = "⛔ BLOCKED" if b+nr+np_==len(ct) else ("❌ FAILING" if f>0 else ("✅ PARTIAL" if p>0 else "— PENDING"))
        row_data = [cat, len(ct), p, f, b, nr, status_lbl]
        r = 13 + i
        bg = GRAY_ROW1 if i%2==0 else GRAY_ROW2
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = font("Arial", 10, False, BRAND_BLACK)
            cell.fill = fill(bg)
            cell.alignment = align("center","center")
            cell.border = thin_border()
        ws.row_dimensions[r].height = 20

    # ── Findings ──
    findings_row = 13 + len(CAT_ORDER) + 1
    ws.row_dimensions[findings_row].height = 14

    ws.merge_cells(f"A{findings_row+1}:G{findings_row+1}")
    ws[f"A{findings_row+1}"] = "KEY FINDINGS & BLOCKERS"
    ws[f"A{findings_row+1}"].font = font("Arial", 11, True, WHITE)
    ws[f"A{findings_row+1}"].fill = fill(BRAND_GOLD)
    ws[f"A{findings_row+1}"].alignment = align("left","center")
    ws.row_dimensions[findings_row+1].height = 24

    findings = [
        ("❌", "No application source code exists — project is pre-scaffolding phase (no src/ directory)", FAIL_FG),
        ("✅", "node_modules installed, postcss.config.js, .env.example, .claude/PLAN.md present", PASS_FG),
        ("⚠️", "schema.prisma missing — database models not defined yet", BLOCKED_FG),
        ("⚠️", "NEXTAUTH_URL=http://localhost in .env.example — MUST be https:// in production", BLOCKED_FG),
        ("✅", "bcryptjs installed — password hashing ready by design", PASS_FG),
        ("✅", "Prisma ORM prevents SQL injection by design; Next.js escapes XSS by default", PASS_FG),
        ("✅", "Payment stack (COD + bKash + Nagad) avoids storing raw card data — PCI scope is minimal", PASS_FG),
        ("❌", "TC-028 Multi-language marked NOT PLANNED — single market (Bangladesh)", NOTRUN_FG),
    ]

    for j, (icon, text, fg_color) in enumerate(findings):
        r = findings_row + 2 + j
        ws.merge_cells(f"B{r}:G{r}")
        ws[f"A{r}"] = icon
        ws[f"A{r}"].alignment = align("center","center")
        ws[f"B{r}"] = text
        ws[f"B{r}"].font = font("Arial", 10, False, fg_color)
        ws[f"B{r}"].fill = fill(GRAY_ROW1 if j%2==0 else GRAY_ROW2)
        ws[f"B{r}"].alignment = align("left","center",True)
        ws.row_dimensions[r].height = 20

    next_row = findings_row + 2 + len(findings) + 1
    ws.merge_cells(f"A{next_row}:G{next_row}")
    ws[f"A{next_row}"] = "⚡ NEXT STEP: Run  npx create-next-app@latest  in E:\\Togor_and_Tweed  →  then see Action Plan sheet"
    ws[f"A{next_row}"].font = font("Arial", 11, True, BRAND_BLACK)
    ws[f"A{next_row}"].fill = fill(LIGHT_GOLD)
    ws[f"A{next_row}"].alignment = align("center","center")
    ws.row_dimensions[next_row].height = 28

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 18


# ─────────────────────────────────────────────────────────────────────────────
def build_test_cases(wb):
    ws = wb.create_sheet("Test Cases")
    ws.freeze_panes = "A2"

    headers = ["Test ID","Category","Scenario","Test Steps","Expected Result","Status","Priority","Actual Result / Notes","QA Date"]
    widths  = [10, 14, 26, 32, 30, 13, 10, 64, 12]

    # Header row
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = font("Arial", 10, True, WHITE)
        cell.fill = fill(BRAND_BLACK)
        cell.alignment = align("center","center",True)
        cell.border = medium_border_bottom()
        ws.column_dimensions[col_letter(ci)].width = w
    ws.row_dimensions[1].height = 28

    # Data rows
    for ri, t in enumerate(TESTS, 2):
        bg = GRAY_ROW1 if ri%2==0 else GRAY_ROW2
        sc = STATUS_STYLE.get(t[5], STATUS_STYLE["NOT RUN"])
        pc = PRIO_BG.get(t[6], WHITE)
        row_vals = [t[0], t[1], t[2], t[3], t[4], t[5], t[6], t[7], QA_DATE]

        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            is_status = ci == 6
            is_prio   = ci == 7
            cell.font = font("Arial", 9, is_status, sc[1] if is_status else BRAND_BLACK)
            cell.fill = fill(sc[0] if is_status else (pc if is_prio else bg))
            cell.alignment = align("left" if ci == 8 else "center", "center", ci==8)
            cell.border = thin_border("E0E0E0")
        ws.row_dimensions[ri].height = 36


# ─────────────────────────────────────────────────────────────────────────────
def build_by_category(wb):
    ws = wb.create_sheet("By Category")
    headers = ["Category","Total","Pass","Fail","Blocked","Not Run","Not Planned","Pass %","Status"]
    widths  = [16, 8, 8, 8, 10, 10, 14, 12, 18]

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = font("Arial", 10, True, WHITE)
        cell.fill = fill(BRAND_BLACK)
        cell.alignment = align("center","center")
        cell.border = medium_border_bottom()
        ws.column_dimensions[col_letter(ci)].width = w
    ws.row_dimensions[1].height = 24

    for ri, cat in enumerate(CAT_ORDER, 2):
        ct = [t for t in TESTS if t[1] == cat]
        if not ct: continue
        p  = sum(1 for t in ct if t[5]=="PASS")
        f  = sum(1 for t in ct if t[5]=="FAIL")
        b  = sum(1 for t in ct if t[5]=="BLOCKED")
        nr = sum(1 for t in ct if t[5]=="NOT RUN")
        np_= sum(1 for t in ct if t[5]=="NOT PLANNED")
        pct= f"{round(p/len(ct)*100)}%" if ct else "0%"
        st = "⛔ BLOCKED" if b+nr+np_==len(ct) else ("❌ FAILING" if f>0 else ("✅ PARTIAL" if p>0 else "— PENDING"))
        row_data = [cat, len(ct), p, f, b, nr, np_, pct, st]
        bg = GRAY_ROW1 if ri%2==0 else GRAY_ROW2
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = font("Arial", 10)
            cell.fill = fill(bg)
            cell.alignment = align("center","center")
            cell.border = thin_border("E0E0E0")
        ws.row_dimensions[ri].height = 20

    # Totals
    tr = len(CAT_ORDER) + 2
    total   = len(TESTS)
    totals  = [
        "TOTAL", total,
        sum(1 for t in TESTS if t[5]=="PASS"),
        sum(1 for t in TESTS if t[5]=="FAIL"),
        sum(1 for t in TESTS if t[5]=="BLOCKED"),
        sum(1 for t in TESTS if t[5]=="NOT RUN"),
        sum(1 for t in TESTS if t[5]=="NOT PLANNED"),
        "0%", ""
    ]
    for ci, val in enumerate(totals, 1):
        cell = ws.cell(row=tr, column=ci, value=val)
        cell.font = font("Arial", 10, True, WHITE)
        cell.fill = fill(BRAND_BLACK)
        cell.alignment = align("center","center")
        cell.border = thin_border()
    ws.row_dimensions[tr].height = 24


# ─────────────────────────────────────────────────────────────────────────────
def build_action_plan(wb):
    ws = wb.create_sheet("Action Plan")

    ws.merge_cells("A1:E1")
    ws["A1"] = "TORGOR & TWEED — QA UNBLOCK ACTION PLAN"
    ws["A1"].font = font("Arial", 14, True, WHITE)
    ws["A1"].fill = fill(BRAND_BLACK)
    ws["A1"].alignment = align("center","center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:E2")
    ws["A2"] = f"Generated: {QA_DATE}  |  Complete these steps in order to enable testing"
    ws["A2"].font = font("Arial", 10, False, "555555", italic=True)
    ws["A2"].fill = fill(LIGHT_GOLD)
    ws["A2"].alignment = align("center","center")
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 10

    headers = ["Phase", "Action Required", "Unblocks Test Cases", "Priority", "Est. Effort"]
    widths  = [20, 60, 30, 12, 16]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = font("Arial", 10, True, WHITE)
        cell.fill = fill(BRAND_GOLD)
        cell.alignment = align("center","center")
        cell.border = thin_border()
        ws.column_dimensions[col_letter(ci)].width = w
    ws.row_dimensions[4].height = 24

    actions = [
        ("1 – Scaffold", r"Run: npx create-next-app@latest in E:\Togor_and_Tweed", "All 54 tests", "Critical", "30 min"),
        ("1 – Scaffold", "Create tsconfig.json, tailwind.config.ts, next.config.ts", "All 54 tests", "Critical", "30 min"),
        ("2 – Database", "Create prisma/schema.prisma (User, Product, Order, Cart, Review, Wishlist, Coupon)", "TC-001–020, TC-033, TC-040, TC-052", "Critical", "2–3 hrs"),
        ("2 – Database", "Run: npx prisma migrate dev --name init", "TC-001–020", "Critical", "15 min"),
        ("3 – Auth",     "Create src/lib/auth.ts with NextAuth v5 + PrismaAdapter + bcryptjs", "TC-013–016, TC-040, TC-041", "High", "2 hrs"),
        ("3 – Auth",     "Set NEXTAUTH_SECRET and GOOGLE_CLIENT_ID/SECRET in .env", "TC-013–016", "High", "30 min"),
        ("3 – Auth",     "Change NEXTAUTH_URL to https://yourdomain.com in production .env", "TC-038", "Critical", "5 min"),
        ("4 – Core Pages","Build: Homepage (brand hero, featured products, categories)", "TC-021, TC-022, TC-023, TC-029", "High", "4–6 hrs"),
        ("4 – Core Pages","Build: Product Listing + Filters + Sorting", "TC-001–003, TC-024, TC-025", "High", "6–8 hrs"),
        ("4 – Core Pages","Build: Product Detail Page (images, description, reviews)", "TC-004, TC-019", "High", "4 hrs"),
        ("4 – Core Pages","Build: Cart (Zustand store + add/remove/quantity)", "TC-005–007, TC-050", "High", "4–6 hrs"),
        ("4 – Core Pages","Build: Checkout (COD flow + order model + confirmation)", "TC-008–009, TC-011, TC-051", "High", "6–8 hrs"),
        ("4 – Core Pages","Build: Auth pages (register, login, logout, reset password)", "TC-013–016", "High", "4 hrs"),
        ("4 – Core Pages","Build: Account Dashboard (My Orders, Wishlist)", "TC-012, TC-017", "Medium", "4 hrs"),
        ("5 – Payments", "Integrate bKash API (checkout.pay.bka.sh)", "TC-010", "Medium", "1 day"),
        ("5 – Payments", "Integrate Nagad API (api.mynagad.com)", "TC-010", "Medium", "1 day"),
        ("5 – Email",    "Configure Resend SDK — order confirmation + welcome emails", "TC-011, TC-013, TC-016", "Medium", "2–4 hrs"),
        ("6 – Features", "Coupon code system (Zod validation)", "TC-020, TC-026, TC-027", "Medium", "4 hrs"),
        ("6 – Features", "Product compare & reviews", "TC-018, TC-019", "Low", "4 hrs"),
        ("7 – Deploy",   "Deploy to Hostinger VPS — PM2 + Nginx + certbot SSL", "TC-029–034, TC-038, TC-042–049", "High", "4–6 hrs"),
        ("7 – Deploy",   "Connect Cloudinary CDN for product images", "TC-034", "Medium", "1–2 hrs"),
        ("8 – QA Run",   "Full regression run — all 54 test cases against live site", "All 54", "Critical", "1 day"),
    ]

    for ri, (phase, action, unblocks, prio, effort) in enumerate(actions, 5):
        bg = GRAY_ROW1 if ri%2==0 else GRAY_ROW2
        pc = {"Critical":PRIO_BG["Critical"],"High":PRIO_BG["High"],"Medium":PRIO_BG["Medium"],"Low":PRIO_BG["Low"]}.get(prio, WHITE)
        row_data = [phase, action, unblocks, prio, effort]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = font("Arial", 10, ci==1, BRAND_BLACK)
            cell.fill = fill(pc if ci==4 else bg)
            cell.alignment = align("left","center", ci==2)
            cell.border = thin_border("E0E0E0")
        ws.row_dimensions[ri].height = 24


# ─────────────────────────────────────────────────────────────────────────────
def main():
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    build_dashboard(wb)
    build_test_cases(wb)
    build_by_category(wb)
    build_action_plan(wb)

    wb.save(OUTPUT)
    print(f"✅  QA Tracker saved: {OUTPUT}")
    print(f"    Sheets: {', '.join(wb.sheetnames)}")
    print(f"    Test cases: {len(TESTS)}")
    print(f"    Blocked: {count_by(5, 'BLOCKED')}  |  Not Run: {count_by(5, 'NOT RUN')}  |  Not Planned: {count_by(5, 'NOT PLANNED')}")
    print(f"    Pass rate: 0% (pre-scaffolding phase)")

if __name__ == "__main__":
    main()
